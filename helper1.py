import json
import os
import sys
import time
from datetime import datetime
from openpyxl import load_workbook
import serial
from time import sleep as wait

# Get current time
current_datetime = datetime.now()

# -------------------- Path Utilities --------------------

def get_data_directory():
    """
    Determine the base path of the data folder depending on whether
    the script is running as a bundled executable or a Python script.
    """
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_dir, "Data")

# Set important folder paths
data_directory = get_data_directory() + "\\"
attendance_folder = f"{data_directory}Attendence Folder\\"
json_database_file = f"{data_directory}Template\\Db.json"
icon_file = f"{data_directory}Template\\icon.ico"
# -------------------- Student Data Utilities --------------------

def get_student_details(tag_id):
    """
    Retrieve student details from the JSON database based on RFID tag ID.
    Returns student dictionary if found, else "no data".
    """
    with open(json_database_file, "r") as file:
        student_data = json.load(file)
        
    if tag_id in student_data:
        return student_data[tag_id]
    else:
        return "no data"

# -------------------- Attendance Management --------------------

def update_attendance(tag_id):
    """
    Mark attendance for the student with the given tag_id.
    Returns status message and student name (if available).
    """
    student = get_student_details(tag_id)
    if student == "no data":
        print("No Such Entry Available")
        return "Not Registered", None

    student_class = student['class']
    current_month = current_datetime.strftime("%B")
    current_day = current_datetime.day

    # Load workbook and sheet for initial processing
    workbook_path = f"{attendance_folder}{student_class}.xlsx"
    wb = load_workbook(workbook_path)
    sheet_april = wb["April"]  # Initial sheet for row lookup
    target_row = None

    # Find the row of the student using tag_id
    for row in sheet_april.iter_rows(min_row=1, max_row=sheet_april.max_row, min_col=1, max_col=1):
        if row[0].value == tag_id:
            target_row = row[0].row
            break

    # Find the column for current date
    target_col = None
    for cell in sheet_april[1]:
        if cell.value == current_day:
            target_col = cell.column_letter
            break

    if target_row is None or target_col is None:
        return "Error", None

    # Construct cell address
    attendance_cell = f"{target_col}{target_row}"

    # Load the actual sheet for current month to update attendance
    wb = load_workbook(workbook_path)
    current_month_sheet = wb[current_month]

    # Get student name for feedback
    student_name = (sheet_april[f"C{target_row}"].value).split(" ")[0]

    # Check if already marked
    if current_month_sheet[attendance_cell].value:
        return "Already Punched", student_name

    # Mark attendance with timestamp
    punch_time = datetime.now().strftime("%I:%M:%S %p")
    current_month_sheet[attendance_cell] = f"P - {punch_time}"
    wb.save(workbook_path)

    return "Marked Present", student_name

# -------------------- Serial Communication --------------------

def connect_arduino(com_port):
    """
    Establish connection with Arduino via the given COM port.
    Sends initialization command and waits for a response.
    Returns (True, SerialObj) if successful, else (False, None).
    """
    try:
        serial_obj = serial.Serial(com_port, 115200, timeout=1, write_timeout=2)
        wait(2)  # Allow time for connection
    except:
        return False, None

    try:
        serial_obj.write("Initialize\n".encode())
    except:
        return False, None

    start_time = time.time()
    timeout_duration = 1  # seconds

    while True:
        if serial_obj.in_waiting > 0:
            response = serial_obj.readline().decode('utf-8').strip()
            if "present" in response:
                return True, serial_obj
        if time.time() - start_time > timeout_duration:
            serial_obj.close()
            return False, None

# -------------------- Main Serial Listening Loop --------------------

def listen_for_cards(serial_obj):
    """
    Continuously read RFID input from Arduino and process attendance.
    Sends feedback messages back to Arduino display.
    """
    display_timer = 0

    try:
        while True:
            if serial_obj.in_waiting > 0:
                data = serial_obj.readline().decode('utf-8').strip()

                if "UID tag:" in data:
                    try:
                        feedback, student_name = update_attendance(data.replace("UID tag: ", "")) # Passing the tag id
                    except Exception as e:
                        print(f"Error updating attendance: {e}")
                        return "Error in Updating Attendance"
                    
                    if feedback == "Not Registered":
                        message = "display Not Registered"
                    elif feedback == "Already Punched":
                        message = f"display {student_name},Already Punched"
                    else:
                        message = f"display {student_name},Marked Present"
                    serial_obj.write(message.encode())
                    display_timer = datetime.now()

            # Reset Arduino display message after 5 seconds
            if display_timer:
                elapsed = (datetime.now() - display_timer).total_seconds()
                if elapsed >= 5:
                    display_timer = 0
                    serial_obj.write("display ".encode())
    except:
        return False

# -------------------- End of Script --------------------


