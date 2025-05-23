import json
from openpyxl import load_workbook
import serial
from time import sleep as wait
import time
import os
import sys

# Get path to the "Data" directory depending on whether it's a script or compiled exe
def get_data_path():
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, "Data")

# Define all relevant paths
data_path = get_data_path()
register_dir = os.path.join(data_path, "Attendence Folder")
template_file_path = os.path.join(data_path, "Template", "template.xlsx")
json_db_path = os.path.join(data_path, "Template", "Db.json")
icon_file = os.path.join(data_path, "Template", "icon.ico")

# Function to read UID from Arduino via Serial
def get_uid_from_arduino(com_port):
    try:
        serial_port = serial.Serial(com_port, 115200, timeout=2, write_timeout=2)
        wait(2)  # Give time for serial port to establish
        connection_successful = True
    except:
        return False, None

    # Initial handshake
    try:
        serial_port.write("Initialize\n".encode())
        time.sleep(0.1)
    except:
        return False, None

    # Wait for confirmation from Arduino
    start_time = time.time()
    timeout_seconds = 1
    while True:
        if serial_port.in_waiting > 0:
            response = serial_port.readline().decode('utf-8').strip()
            if "present" in response:
                break
        if time.time() - start_time > timeout_seconds:
            serial_port.close()
            print("stopped")
            return False, None

    # Request UID from Arduino
    serial_port.write("display Scan Ur Card:- ".encode())
    start_time = time.time()
    timeout_seconds = 5
    while True:
        if serial_port.in_waiting > 0:
            response = serial_port.readline().decode('utf-8').strip()
            if "UID tag:" in response:
                break
        if time.time() - start_time > timeout_seconds:
            serial_port.close()
            return True, "retry"

    # Send final message and parse UID
    serial_port.write("display ".encode())
    uid = response.replace("UID tag: ", "")
    print(uid)
    return connection_successful, str(uid)

# Create new class register Excel file from template
def create_class_register(class_name):
    file_path = os.path.join(register_dir, class_name)
    print(file_path)
    if os.path.exists(file_path):
        return "Class's Register already exists"
    try:
        workbook = load_workbook(template_file_path)
    except Exception as e:
        print(e)
        return "Error Loading Template"
    workbook.save(os.path.join(register_dir, class_name))
    print("Data saved using the Excel template.")
    return "Class's Register Created"

# Add new student using their RFID
def add_student(class_name, roll_number, student_name, com_port):
    success, tag_id = get_uid_from_arduino(com_port)
    if not success:
        print("Wrong Port Selected")
        return "Wrong Port Selected"
    if "retry" in tag_id:
        return "No Card Detected"

    with open(json_db_path, "r") as f:
        db_data = json.load(f)

    if tag_id in db_data:
        print("Tag_Id Already Defined")
        return "Tag Already Defined"

    for entry in db_data.values():
        if entry["name"] == student_name and entry["roll_no"] == roll_number:
            print("Student with same name and roll number already exists.")
            return "Student Already Exists"

    workbook = load_workbook(os.path.join(register_dir, class_name + ".xlsx"))
    worksheet = workbook["April"]
    worksheet.append([tag_id, roll_number, student_name])
    workbook.save(os.path.join(register_dir, class_name + ".xlsx"))

    db_data[tag_id] = {
        "class": class_name,
        "roll_no": roll_number,
        "name": student_name
    }
    with open(json_db_path, "w") as f:
        json.dump(db_data, f, indent=4)

    print(f"New student added with tag_id {tag_id} and file updated successfully.")
    return "New Student Added Successfully"

# Read student data from scanned RFID
def extract_student_data(com_port):
    success, tag_id = get_uid_from_arduino(com_port)
    if not success:
        print("Wrong Port Selected")
        return None, None, None, "Wrong Port Selected"
    if "retry" in tag_id:
        return None, None, None, "No Card Detected"

    with open(json_db_path, "r") as f:
        db_data = json.load(f)

    if tag_id not in db_data:
        print("Tag ID not found.")
        return None, None, None, "Tag ID Not Found"

    student = db_data[tag_id]
    print(f"Name     : {student['name']}")
    print(f"Class    : {student['class']}")
    print(f"Roll No. : {student['roll_no']}")
    return student['name'], student['class'], student['roll_no'], None

# Update details of a registered student using RFID
def update_student_details(class_name, roll_number, student_name, com_port):
    success, tag_id = get_uid_from_arduino(com_port)
    if not success:
        print("Wrong Port Selected")
        return "Wrong Port Selected"
    if "retry" in tag_id:
        return "No Card Detected"

    with open(json_db_path, "r") as f:
        db_data = json.load(f)

    if tag_id not in db_data:
        print("This Card is not registered.")
        return

    db_data[tag_id] = {
        "class": class_name,
        "roll_no": roll_number,
        "name": student_name
    }

    with open(json_db_path, "w") as f:
        json.dump(db_data, f, indent=4)

    workbook = load_workbook(os.path.join(register_dir, class_name + ".xlsx"))
    worksheet = workbook["April"]
    found = False
    for row in worksheet.iter_rows(min_row=2):
        if str(row[0].value) == tag_id:
            row[1].value = roll_number
            row[2].value = student_name
            found = True
            break
    workbook.save(os.path.join(register_dir, class_name + ".xlsx"))

    if not found:
        return "No Data Found In Register"
    return "Value Updated Succesfully"

# Replace an old RFID with a new one
def change_rfid_tag(class_name, roll_number, student_name, com_port):
    success, new_tag_id = get_uid_from_arduino(com_port)
    if not success:
        print("Wrong Port Selected")
        return "Wrong Port Selected"
    if "retry" in new_tag_id:
        return "No Card Detected"

    with open(json_db_path, "r") as f:
        db_data = json.load(f)

    if new_tag_id in db_data:
        return "Tag Already Defined"

    old_tag_id = None
    for tag, info in db_data.items():
        if info["class"] == class_name and info["roll_no"] == roll_number and info["name"] == student_name:
            old_tag_id = tag
            break

    if old_tag_id is None:
        print("Student not found in Database.")
        return "Student not found in Database"

    db_data[new_tag_id] = db_data.pop(old_tag_id)
    with open(json_db_path, "w") as f:
        json.dump(db_data, f, indent=4)
    print(f"Tag ID changed in JSON: {old_tag_id} -> {new_tag_id}")

    workbook = load_workbook(os.path.join(register_dir, class_name + ".xlsx"))
    worksheet = workbook["April"]
    found = False
    for row in worksheet.iter_rows(min_row=2):
        if str(row[0].value) == old_tag_id:
            row[0].value = new_tag_id
            found = True
            print("Tag ID updated in Excel.")
            break
    workbook.save(os.path.join(register_dir, class_name + ".xlsx"))

    if not found:
        return "Tag Id Not Found in Register"
    return "Tag ID Changed Successfully"

# Remove a student using their tag ID
def delete_tag(tag_id):
    with open(json_db_path, "r") as f:
        db_data = json.load(f)

    if tag_id not in db_data:
        print(f"ID {tag_id} not found in JSON.")
        return False

    class_name = db_data[tag_id]['class']
    del db_data[tag_id]
    with open(json_db_path, "w") as f:
        json.dump(db_data, f, indent=4)
    print(f"Deleted ID: {tag_id} from JSON.")

    workbook = load_workbook(os.path.join(register_dir, class_name + ".xlsx"))
    worksheet = workbook["April"]
    row_to_delete = None
    for row in worksheet.iter_rows(min_row=1, max_col=1):
        if row[0].value == tag_id:
            row_to_delete = row[0].row
            break

    if row_to_delete:
        worksheet.delete_rows(row_to_delete)
        workbook.save(os.path.join(register_dir, class_name + ".xlsx"))
        print(f"Deleted row {row_to_delete} in Excel for ID: {tag_id}")
    else:
        print(f"ID {tag_id} not found in Excel sheet.")
    return True

if __name__ == '__main__':
    # Uncomment for testing:
    # get_uid_from_arduino("COM3")
    # delete_tag("1c5d9282")
    pass
